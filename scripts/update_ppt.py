"""
PPT 임베드 Excel 데이터 업데이트 스크립트

입력:  input data/SK electlink '26.4월 경영회의_성장지원팀.v1.pptx
출력:  input data/SK electlink '26.4월 경영회의_성장지원팀.v2.pptx

처리:
- 10개 임베드 xlsx에 Mar-26 데이터 추가
- chart XML 캐시 업데이트
- "Epic" → "SK일렉링크" 레이블 변경

Usage:
    python update_ppt.py
    python update_ppt.py --dry-run   # xlsx만 출력, PPTX 저장 안 함
"""

import argparse
import io
import json
import re
import zipfile
from datetime import date
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
INPUT_DATA = REPO_ROOT.parent / "input data"
SRC_PPTX = INPUT_DATA / "SK electlink '26.4월 경영회의_성장지원팀.v1.pptx"
DST_PPTX = INPUT_DATA / "SK electlink '26.4월 경영회의_성장지원팀.v2.pptx"
MONTHLY_JSON = REPO_ROOT / "data" / "monthly" / "2026-03.json"


def excel_serial(d: date) -> int:
    """Python date → Excel 시리얼 날짜 (1900 윤년 버그 포함)."""
    return (d - date(1899, 12, 30)).days


MAR26 = date(2026, 3, 25)
MAR26_SERIAL = excel_serial(MAR26)


def load_monthly() -> dict:
    with open(MONTHLY_JSON, encoding="utf-8") as f:
        return json.load(f)


def build_mar26_data(m: dict) -> dict:
    """2026-03.json에서 Mar-26 업데이트 값 계산."""
    fast_total = m["fast"]["total"]
    slow_total = m["slow"]["total"]
    fc = m["fast_concentration"]
    fr = m["fast_regional"]

    return {
        # chart1: Fast Overview (sliding 5)
        "chart1": {"date": MAR26, "vals": [fast_total, None]},
        # chart2: Slow Overview (sliding 5)
        "chart2": {"date": MAR26, "vals": [slow_total, None]},
        # chart3: Fast M/S nationwide (horizontal, sliding 5)
        "chart3": {
            "label": "Mar-26",
            "by_series": {
                "DY Chaevi": round(fr["전국"]["operators"]["채비"]["ms_pct"] / 100, 5),
                "Epic":      round(fr["전국"]["operators"]["SK일렉링크"]["ms_pct"] / 100, 5),
                "Evsis":     round(fr["전국"]["operators"]["이브이시스"]["ms_pct"] / 100, 5),
                "Humax + Jeju Electric": round(1782 / fast_total, 5),
                "GS Chargev":            round(1604 / fast_total, 5),
            },
        },
        # chart4: Slow M/S (horizontal, sliding 5)
        "chart4": {
            "label": "Mar-26",
            "by_series": {
                "GS Connect + ChargEV": round(77317 / slow_total, 5),
                "PowerCube":            round(69072 / slow_total, 5),
                "EverOn":               round(50544 / slow_total, 5),
                "Volt-up":              round(40184 / slow_total, 5),
                "Pluglink":             round(42132 / slow_total, 5),
            },
        },
        # chart5: Slow Trend counts (full history, vertical)
        "chart5": {
            "date": MAR26,
            "vals": [77317, 69072, 50544, 40184, 42132, 4963],
            # cols: GS ChargEV, PowerCube, EverOn, Volt-up, Pluglink, Epic→SK
        },
        # chart6: Fast Trend counts (full history, vertical)
        "chart6": {
            "date": MAR26,
            "vals": [5916, 5089, 2571, 1782, 1604],
            # cols: Chaevi, Epic→SK, EVSIS, Humax+JES, GSChargEV
        },
        # chart7: Concentration GSMA+Metro (full history, vertical)
        "chart7": {
            "date": MAR26,
            "vals": [
                round(fc["채비"]["GSMA_plus_metro_pct"] / 100, 4),
                round(fc["SK일렉링크"]["GSMA_plus_metro_pct"] / 100, 4),
                round(fc["이브이시스"]["GSMA_plus_metro_pct"] / 100, 4),
            ],
        },
        # chart8: Concentration GSMA (full history, vertical)
        "chart8": {
            "date": MAR26,
            "vals": [
                round(fc["채비"]["GSMA_pct"] / 100, 4),
                round(fc["SK일렉링크"]["GSMA_pct"] / 100, 4),
                round(fc["이브이시스"]["GSMA_pct"] / 100, 4),
            ],
        },
        # chart9: M/S GSMA (full history, vertical)
        "chart9": {
            "date": MAR26,
            "vals": [
                round(fr["GSMA"]["operators"]["채비"]["ms_pct"] / 100, 5),
                round(fr["GSMA"]["operators"]["SK일렉링크"]["ms_pct"] / 100, 5),
                round(fr["GSMA"]["operators"]["이브이시스"]["ms_pct"] / 100, 5),
            ],
        },
        # chart10: M/S GSMA+광역시 (full history, vertical)
        "chart10": {
            "date": MAR26,
            "vals": [
                round(fr["GSMA+광역시"]["operators"]["채비"]["ms_pct"] / 100, 5),
                round(fr["GSMA+광역시"]["operators"]["SK일렉링크"]["ms_pct"] / 100, 5),
                round(fr["GSMA+광역시"]["operators"]["이브이시스"]["ms_pct"] / 100, 5),
            ],
        },
    }


# ─────────────────── xlsx 업데이트 ───────────────────

def rename_epic_xlsx(ws):
    """첫 행 또는 첫 열에서 'Epic' → 'SK일렉링크' 교체."""
    for cell in list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]:
        if cell.value == "Epic":
            cell.value = "SK일렉링크"
    for cell in list(ws.iter_cols(min_col=1, max_col=1, values_only=False))[0]:
        if cell.value == "Epic":
            cell.value = "SK일렉링크"


def update_xlsx_vertical_full(ws, new_date: date, new_vals: list) -> int:
    """
    날짜=행 전체이력 차트: 마지막 데이터 행 다음에 Mar-26 추가.
    Returns: 새로 추가된 행 번호.
    """
    last_row = 1
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is not None:
            last_row = row[0].row
    new_row = last_row + 1
    ws.cell(row=new_row, column=1).value = new_date
    for i, v in enumerate(new_vals, start=2):
        ws.cell(row=new_row, column=i).value = v
    return new_row


def update_xlsx_vertical_sliding(ws, new_date: date, new_vals: list) -> int:
    """
    날짜=행 슬라이딩 윈도우 차트 (chart1, chart2): 마지막 행 다음에 Mar-26 추가.
    Returns: 새로 추가된 행 번호.
    """
    return update_xlsx_vertical_full(ws, new_date, new_vals)


def update_xlsx_horizontal_sliding(ws, new_label: str, by_series: dict):
    """
    날짜=열 슬라이딩 윈도우 차트 (chart3, chart4): 마지막 열 다음에 Mar-26 추가.
    Returns: 새로 추가된 열 번호.
    """
    # 헤더 행(1행)에서 마지막 데이터 열 찾기
    last_col = 1
    for cell in ws[1]:
        if cell.value is not None and cell.value != " ":
            last_col = cell.column
    new_col = last_col + 1

    # 헤더 추가
    ws.cell(row=1, column=new_col).value = new_label

    # 각 시리즈 행에 값 추가 (Epic이 SK일렉링크로 이미 변경된 경우도 처리)
    lookup = {**by_series, "SK일렉링크": by_series.get("Epic", by_series.get("SK일렉링크"))}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        series_name = row[0].value
        if series_name in lookup and lookup[series_name] is not None:
            row[new_col - 1].value = lookup[series_name]

    return new_col


# ─────────────────── chart XML 업데이트 ───────────────────

def rename_epic_xml(xml: str) -> str:
    """chart XML에서 series명 Epic → SK일렉링크 교체."""
    # <c:tx> 안의 <c:v>Epic</c:v>
    return re.sub(r'(<c:tx>.*?<c:v>)Epic(</c:v>)', r'\1SK일렉링크\2', xml, flags=re.DOTALL)


def col_letter(n: int) -> str:
    """1-indexed 열 번호 → 엑셀 열 문자 (1=A, 27=AA, ...)."""
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def update_ref_range(ref: str, row_delta: int = 0, col_delta: int = 0, extend: bool = True) -> str:
    """
    Excel 범위 참조 업데이트.
    - extend=True: 마지막 행/열을 +1 (full history)
    - extend=False: 전체 윈도우를 +1 (sliding window)
    """
    # 패턴: Sheet1!$A$14:$A$39
    m = re.match(r"(.*!\$)([A-Z]+)(\$)(\d+)(:?\$?)([A-Z]*)(\$?)(\d*)", ref)
    if not m:
        return ref

    sheet_part = m.group(1)
    col1 = m.group(2)
    row1 = int(m.group(4))

    if ":" not in ref:
        return ref

    col2 = m.group(6)
    row2 = int(m.group(8)) if m.group(8) else row1

    if row_delta:
        if extend:
            row2 += row_delta
        else:
            row1 += row_delta
            row2 += row_delta
    if col_delta:
        col1_n = sum((ord(c) - 64) * (26 ** i) for i, c in enumerate(reversed(col1)))
        col2_n = sum((ord(c) - 64) * (26 ** i) for i, c in enumerate(reversed(col2)))
        if extend:
            col2_n += col_delta
            col2 = col_letter(col2_n)
        else:
            col1_n += col_delta
            col2_n += col_delta
            col1 = col_letter(col1_n)
            col2 = col_letter(col2_n)

    return f"{sheet_part}{col1}${row1}:${col2}${row2}"


def xml_add_pt_to_numcache(block: str, new_val, pt_count: int) -> str:
    """numCache 블록에 새 pt 추가 + ptCount 증가."""
    # ptCount 증가
    block = re.sub(
        r'(<c:ptCount val=")(\d+)(")',
        lambda m: f'{m.group(1)}{int(m.group(2)) + 1}{m.group(3)}',
        block, count=1
    )
    # 마지막 </c:numCache> 직전에 새 pt 삽입
    val_str = "" if new_val is None else str(new_val)
    new_pt = f'<c:pt idx="{pt_count}"><c:v>{val_str}</c:v></c:pt>'
    block = block.replace("</c:numCache>", f"{new_pt}</c:numCache>", 1)
    return block


def xml_add_pt_to_strcache(block: str, new_val: str, pt_count: int) -> str:
    """strCache 블록에 새 pt 추가 + ptCount 증가."""
    block = re.sub(
        r'(<c:ptCount val=")(\d+)(")',
        lambda m: f'{m.group(1)}{int(m.group(2)) + 1}{m.group(3)}',
        block, count=1
    )
    new_pt = f'<c:pt idx="{pt_count}"><c:v>{new_val}</c:v></c:pt>'
    block = block.replace("</c:strCache>", f"{new_pt}</c:strCache>", 1)
    return block


def xml_shift_pt_window(block: str, new_val, new_label: str = None) -> str:
    """
    슬라이딩 윈도우: 기존 5개 pt 중 첫 번째 제거 후 새 값 추가 (idx 재정렬).
    numCache 또는 strCache 둘 다 처리.
    """
    is_str = "<c:strCache>" in block

    if is_str:
        pts = re.findall(r'<c:pt idx="\d+"><c:v>(.*?)</c:v></c:pt>', block)
        pts = pts[1:] + [new_label or str(new_val)]  # 첫 번째 제거, 새 값 추가
        new_pts = "\n".join(
            f'<c:pt idx="{i}"><c:v>{v}</c:v></c:pt>' for i, v in enumerate(pts)
        )
        block = re.sub(
            r'<c:pt idx="\d+">.*?</c:pt>',
            lambda m, used=[False]: "" if not used[0] and not used.__setitem__(0, True) else m.group(),
            block, flags=re.DOTALL
        )
        # 더 깔끔하게: 전체 pt 블록 교체
        block = re.sub(
            r'(<c:strCache>.*?<c:ptCount[^/]*/>\s*)(.*?)(</c:strCache>)',
            lambda m: m.group(1) + new_pts + m.group(3),
            block, flags=re.DOTALL
        )
    else:
        pts_raw = re.findall(r'<c:pt idx="\d+"><c:v>(.*?)</c:v></c:pt>', block)
        new_v = "" if new_val is None else str(new_val)
        pts = pts_raw[1:] + [new_v]
        new_pts = "\n".join(
            f'<c:pt idx="{i}"><c:v>{v}</c:v></c:pt>' for i, v in enumerate(pts)
        )
        block = re.sub(
            r'(<c:numCache>.*?<c:ptCount[^/]*/>\s*)(.*?)(</c:numCache>)',
            lambda m: m.group(1) + new_pts + m.group(3),
            block, flags=re.DOTALL
        )
    return block


def update_chart_xml_full_history(xml: str, new_date_serial: int, new_vals: list, col_delta: int = 0) -> str:
    """
    전체이력 차트 (chart5-10): 각 series에 새 pt 추가, 범위 참조 확장.
    - col_delta: 0이면 행 확장, 양수면 열 확장 (현재 사용 안 함)
    """
    ser_blocks = re.findall(r'(<c:ser>)(.*?)(</c:ser>)', xml, re.DOTALL)

    # series 순서: 0번은 series 이름(tx), 나머지가 실제 데이터 series
    # val_idx: 실제 데이터 series 인덱스 (name series 제외)
    data_ser_idx = 0
    result = xml

    for match in re.finditer(r'<c:ser>(.*?)</c:ser>', xml, re.DOTALL):
        ser_content = match.group(1)
        tx_match = re.search(r'<c:tx>.*?<c:v>(.*?)</c:v>', ser_content, re.DOTALL)
        if not tx_match:
            continue

        new_ser = ser_content

        # numRef 블록들 처리 (date axis + value)
        numrefs = list(re.finditer(r'<c:numRef>(.*?)</c:numRef>', ser_content, re.DOTALL))

        for nr_idx, nr_match in enumerate(numrefs):
            nr_block = nr_match.group(0)
            inner = nr_match.group(1)

            # ptCount 확인
            ptcount_m = re.search(r'<c:ptCount val="(\d+)"', inner)
            if not ptcount_m:
                continue
            pt_count = int(ptcount_m.group(1))

            if pt_count <= 1:
                continue  # series name block, skip

            # 범위 참조 업데이트 (행 확장)
            new_nr_block = re.sub(
                r'<c:f>(.*?)</c:f>',
                lambda m: f'<c:f>{update_ref_range(m.group(1), row_delta=1, extend=True)}</c:f>',
                nr_block
            )

            # 새 pt 값 결정
            if nr_idx == 0 or "46" in inner[:100]:
                # 날짜 축 (숫자 serial)
                new_v = new_date_serial
            else:
                # 값 축: data_ser_idx 번째 값
                new_v = new_vals[data_ser_idx] if data_ser_idx < len(new_vals) else None

            new_nr_block = xml_add_pt_to_numcache(new_nr_block, new_v, pt_count)
            new_ser = new_ser.replace(nr_match.group(0), new_nr_block)

        data_ser_idx += 1
        result = result.replace(match.group(0), f"<c:ser>{new_ser}</c:ser>", 1)

    return result


def update_chart_xml_sliding_vertical(xml: str, new_date_serial: int, new_vals: list, new_row: int) -> str:
    """
    슬라이딩 윈도우 수직 차트 (chart1, chart2): 범위 참조 이동 + 캐시 값 교체.
    """
    result = xml
    val_idx = 0

    for match in re.finditer(r'<c:ser>(.*?)</c:ser>', xml, re.DOTALL):
        ser_content = match.group(1)
        new_ser = ser_content

        numrefs = list(re.finditer(r'<c:numRef>(.*?)</c:numRef>', ser_content, re.DOTALL))
        for nr_idx, nr_match in enumerate(numrefs):
            nr_block = nr_match.group(0)
            inner = nr_match.group(1)
            ptcount_m = re.search(r'<c:ptCount val="(\d+)"', inner)
            if not ptcount_m:
                continue
            pt_count = int(ptcount_m.group(1))
            if pt_count <= 1:
                continue

            # 범위 참조: 윈도우 전체 +1 이동
            new_nr_block = re.sub(
                r'<c:f>(.*?)</c:f>',
                lambda m: f'<c:f>{update_ref_range(m.group(1), row_delta=1, extend=False)}</c:f>',
                nr_block
            )

            # 새 값 (날짜 or 데이터)
            pts = re.findall(r'<c:pt[^>]*><c:v>(.*?)</c:v></c:pt>', inner)
            is_date_axis = any("46" in p and len(p) >= 5 for p in pts[:3])

            new_v = new_date_serial if is_date_axis else new_vals[val_idx]
            new_nr_block = xml_shift_pt_window(new_nr_block, new_v)
            new_ser = new_ser.replace(nr_match.group(0), new_nr_block)

        val_idx += 1
        result = result.replace(match.group(0), f"<c:ser>{new_ser}</c:ser>", 1)

    return result


def update_chart_xml_sliding_horizontal(xml: str, new_label: str, new_vals_by_name: dict) -> str:
    """
    슬라이딩 윈도우 수평 차트 (chart3, chart4): 열 범위 이동 + 캐시 값 교체.
    """
    result = xml
    val_idx = 0

    for match in re.finditer(r'<c:ser>(.*?)</c:ser>', xml, re.DOTALL):
        ser_content = match.group(1)
        tx_match = re.search(r'<c:tx>.*?<c:v>(.*?)</c:v>', ser_content, re.DOTALL)
        ser_name = tx_match.group(1) if tx_match else ""
        new_ser = ser_content

        # category strRef (month labels) - <c:cat> 안에 있는 것만 업데이트
        # <c:tx> 안의 strRef(시리즈명)는 건드리지 않음
        for cat_match in re.finditer(r'<c:cat>(.*?)</c:cat>', ser_content, re.DOTALL):
            cat_block = cat_match.group(0)
            for str_match in re.finditer(r'<c:strRef>(.*?)</c:strRef>', cat_block, re.DOTALL):
                sr_block = str_match.group(0)
                new_sr = re.sub(
                    r'<c:f>(.*?)</c:f>',
                    lambda m: f'<c:f>{update_ref_range(m.group(1), col_delta=1, extend=False)}</c:f>',
                    sr_block
                )
                new_sr = xml_shift_pt_window(new_sr, None, new_label)
                new_cat = cat_block.replace(str_match.group(0), new_sr)
                new_ser = new_ser.replace(cat_match.group(0), new_cat)

        # numRef (values)
        for nr_match in re.finditer(r'<c:numRef>(.*?)</c:numRef>', ser_content, re.DOTALL):
            nr_block = nr_match.group(0)
            inner = nr_match.group(1)
            ptcount_m = re.search(r'<c:ptCount val="(\d+)"', inner)
            if not ptcount_m:
                continue
            pt_count = int(ptcount_m.group(1))
            if pt_count <= 1:
                continue

            new_nr = re.sub(
                r'<c:f>(.*?)</c:f>',
                lambda m: f'<c:f>{update_ref_range(m.group(1), col_delta=1, extend=False)}</c:f>',
                nr_block
            )
            # Epic이 SK일렉링크로 이미 변경된 경우 fallback
            new_v = new_vals_by_name.get(ser_name) or new_vals_by_name.get("Epic", 0.0)
            new_nr = xml_shift_pt_window(new_nr, new_v)
            new_ser = new_ser.replace(nr_match.group(0), new_nr)

        result = result.replace(match.group(0), f"<c:ser>{new_ser}</c:ser>", 1)

    return result


# ─────────────────── 메인 처리 ───────────────────

CHART_CONFIG = {
    "chart1":  {"xlsx": "Microsoft_Excel_Worksheet.xlsx",  "type": "vertical_sliding"},
    "chart2":  {"xlsx": "Microsoft_Excel_Worksheet1.xlsx", "type": "vertical_sliding"},
    "chart3":  {"xlsx": "Microsoft_Excel_Worksheet2.xlsx", "type": "horizontal_sliding"},
    "chart4":  {"xlsx": "Microsoft_Excel_Worksheet3.xlsx", "type": "horizontal_sliding"},
    "chart5":  {"xlsx": "Microsoft_Excel_Worksheet4.xlsx", "type": "vertical_full",
                "epic_col": 6},   # 열 G = Epic
    "chart6":  {"xlsx": "Microsoft_Excel_Worksheet5.xlsx", "type": "vertical_full",
                "epic_col": 2},   # 열 C = Epic
    "chart7":  {"xlsx": "Microsoft_Excel_Worksheet6.xlsx", "type": "vertical_full",
                "epic_col": 2},
    "chart8":  {"xlsx": "Microsoft_Excel_Worksheet7.xlsx", "type": "vertical_full",
                "epic_col": 2},
    "chart9":  {"xlsx": "Microsoft_Excel_Worksheet8.xlsx", "type": "vertical_full",
                "epic_col": 2},
    "chart10": {"xlsx": "Microsoft_Excel_Worksheet9.xlsx", "type": "vertical_full",
                "epic_col": 2},
}


def process_pptx(data26: dict, dry_run: bool = False):
    monthly = load_monthly()
    mar26 = build_mar26_data(monthly)

    print("\n=== Mar-26 데이터 요약 ===")
    for k, v in mar26.items():
        print(f"  {k}: {v}")

    if dry_run:
        print("\n[DRY-RUN] 실제 파일 저장 없이 종료합니다.")
        return

    with zipfile.ZipFile(SRC_PPTX, "r") as src_zip:
        with zipfile.ZipFile(DST_PPTX, "w", compression=zipfile.ZIP_DEFLATED) as dst_zip:
            # 모든 파일 복사 기반으로 시작
            file_overrides = {}

            # 각 차트 처리
            for chart_name, cfg in CHART_CONFIG.items():
                print(f"\n처리 중: {chart_name} ({cfg['xlsx']})")
                xlsx_path = f"ppt/embeddings/{cfg['xlsx']}"
                chart_xml_path = f"ppt/charts/{chart_name}.xml"

                # xlsx 업데이트
                xlsx_data = src_zip.read(xlsx_path)
                wb = openpyxl.load_workbook(io.BytesIO(xlsx_data))
                ws = wb.active

                rename_epic_xlsx(ws)
                d = mar26[chart_name]
                chart_type = cfg["type"]

                if chart_type == "vertical_full":
                    new_row = update_xlsx_vertical_full(ws, d["date"], d["vals"])
                    print(f"  xlsx: 행 {new_row} 추가 → {d['vals']}")

                elif chart_type == "vertical_sliding":
                    new_row = update_xlsx_vertical_sliding(ws, d["date"], d["vals"])
                    print(f"  xlsx: 행 {new_row} 추가 (슬라이딩 윈도우)")

                elif chart_type == "horizontal_sliding":
                    new_col = update_xlsx_horizontal_sliding(ws, d["label"], d["by_series"])
                    print(f"  xlsx: 열 {new_col} 추가 (Mar-26)")

                buf = io.BytesIO()
                wb.save(buf)
                file_overrides[xlsx_path] = buf.getvalue()

                # chart XML 업데이트
                xml_bytes = src_zip.read(chart_xml_path)
                xml = xml_bytes.decode("utf-8")
                xml = rename_epic_xml(xml)

                if chart_type == "vertical_full":
                    xml = update_chart_xml_full_history(xml, MAR26_SERIAL, d["vals"])
                elif chart_type == "vertical_sliding":
                    xml = update_chart_xml_sliding_vertical(xml, MAR26_SERIAL, d["vals"], new_row)
                elif chart_type == "horizontal_sliding":
                    xml = update_chart_xml_sliding_horizontal(xml, d["label"], d["by_series"])

                file_overrides[chart_xml_path] = xml.encode("utf-8")
                print(f"  chart XML: 캐시 업데이트 완료")

            # 파일 기록
            for item in src_zip.infolist():
                if item.filename in file_overrides:
                    dst_zip.writestr(item, file_overrides[item.filename])
                else:
                    dst_zip.writestr(item, src_zip.read(item.filename))

    print(f"\n저장 완료: {DST_PPTX}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    process_pptx({}, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
